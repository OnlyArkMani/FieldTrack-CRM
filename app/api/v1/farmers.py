"""Farmers (Customer/Farmer DB) router — Module 1. Thin HTTP layer; all logic
and team-scope authorization live in FarmerService.

AUTHZ:
- All endpoints require an authenticated active user. Team scoping is enforced
  inside the service (ADMIN sees all; supervisor/employee see their team).
- Create/update/lead-status are available to any field user for their own
  team's farmers; the service decides what team a new farmer lands in.
"""
import csv
import io
from typing import Annotated

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import CurrentUser, get_current_admin, get_db
from app.core.exceptions import bad_request
from app.models.user import User
from app.schemas.common import CursorPage
from app.schemas.crm import (
    CustomerImportResult,
    FarmerCreate,
    FarmerDetailResponse,
    FarmerListItem,
    FarmerResponse,
    FarmerUpdate,
    LeadHistoryItem,
    LeadResponse,
    LeadStatusUpdate,
    LivestockProfileResponse,
    VisitSummary,
)
from app.services.farmer_service import FarmerService

router = APIRouter(prefix="/farmers", tags=["farmers"])

# Columns of the bulk-import template (order preserved in the .xlsx/.csv).
IMPORT_COLUMNS = [
    "name",
    "customer_type",
    "phone",
    "village",
    "district",
    "address",
    "total_cattle",
    "current_feed_brand",
    "team_id",
    "notes",
]


@router.get("/ping")
async def ping() -> dict:
    return {"status": "ok", "module": "farmers"}


@router.get("", response_model=CursorPage[FarmerListItem])
async def list_farmers(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    cursor: str | None = Query(default=None, description="Opaque forward cursor"),
    limit: int = Query(default=20, ge=1, le=100),
    team_id: int | None = Query(default=None, description="Admin-only team filter"),
    customer_type: str | None = Query(
        default=None,
        description="Filter by type: FARMER_MEET | FPO | VLCC | RETAILER | DISTRIBUTOR",
    ),
    lead_status: str | None = Query(
        default=None, description="Filter by current lead status: HOT | WARM | COLD"
    ),
    search: str | None = Query(
        default=None, max_length=200, description="Match name or village"
    ),
) -> CursorPage[FarmerListItem]:
    """Paginated customer list with the CURRENT lead status joined per row.
    Supervisor/employee see only their team's customers; admin sees all.
    Optional customer_type filter powers the [All][Farmers][FPOs][VLCCs] tabs."""
    ct = customer_type.strip().upper() if customer_type else None
    if ct and ct not in ("FARMER_MEET", "FPO", "VLCC", "RETAILER", "DISTRIBUTOR"):
        raise bad_request(
            "customer_type must be FARMER_MEET, FPO, VLCC, RETAILER or DISTRIBUTOR"
        )
    return await FarmerService(db).list_farmers(
        user=user,
        cursor=cursor,
        limit=limit,
        team_id=team_id,
        lead_status=lead_status.strip().upper() if lead_status else None,
        search=search,
        customer_type=ct,
    )


# ── Bulk import (admin preload) ──────────────────────────────────────────
@router.get("/import/template")
async def import_template(
    admin: Annotated[User, Depends(get_current_admin)],
) -> StreamingResponse:
    """Download the blank customer-import spreadsheet (CSV). Fill it and upload
    it to POST /farmers/import. customer_type must be FARMER_MEET, FPO, VLCC,
    RETAILER or DISTRIBUTOR."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(IMPORT_COLUMNS)
    writer.writerow(
        ["Ramesh Patil", "FARMER_MEET", "9876543210", "Shirur", "Pune", "", "8", "AmulFeed", "", ""]
    )
    writer.writerow(
        ["Shirur Dairy FPO", "FPO", "9876500000", "Shirur", "Pune", "", "", "", "", ""]
    )
    writer.writerow(
        ["Kendur VLCC", "VLCC", "9876511111", "Kendur", "Pune", "", "", "", "", ""]
    )
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=customers_import_template.csv"
        },
    )


@router.post("/import", response_model=CustomerImportResult)
async def import_customers(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
    file: Annotated[UploadFile, File(description="CSV or XLSX, header row required")],
    dry_run: bool = Query(
        default=True,
        description="true = validate and preview only; false = commit the insert",
    ),
) -> CustomerImportResult:
    """Bulk-preload customers. Defaults to a dry run (validation preview);
    pass dry_run=false to actually insert. Admin only."""
    content = await file.read()
    rows = _parse_import_file(file.filename or "", content)
    return await FarmerService(db).import_customers(rows, user=admin, dry_run=dry_run)


def _parse_import_file(filename: str, content: bytes) -> list[dict]:
    """Parse an uploaded CSV or XLSX into a list of row dicts keyed by header."""
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:  # pragma: no cover
            raise bad_request("XLSX import needs openpyxl; upload a CSV instead")
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            return []
        out: list[dict] = []
        for r in rows_iter:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            out.append({header[i]: r[i] if i < len(r) else None for i in range(len(header))})
        return out
    # default: CSV / text
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


@router.get("/{farmer_id}", response_model=FarmerDetailResponse)
async def get_farmer(
    farmer_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> FarmerDetailResponse:
    """Full profile: base info, current lead, last 3 visits, latest livestock,
    pending follow-ups, and total orders/visits."""
    return await FarmerService(db).get_farmer_with_full_profile(farmer_id, user)


@router.post("", response_model=FarmerResponse, status_code=201)
async def create_farmer(
    body: FarmerCreate,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> FarmerResponse:
    """Create a farmer. Employees are pinned to their own team; admin/supervisor
    may set team_id explicitly."""
    return await FarmerService(db).create_farmer(body, user=user)


@router.put("/{farmer_id}", response_model=FarmerResponse)
async def update_farmer(
    farmer_id: int,
    body: FarmerUpdate,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> FarmerResponse:
    """Update base info only (livestock is captured per visit, not here)."""
    return await FarmerService(db).update_farmer(farmer_id, body, user=user)


@router.get("/{farmer_id}/visits", response_model=CursorPage[VisitSummary])
async def farmer_visits(
    farmer_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    cursor: str | None = Query(default=None, description="Opaque forward cursor"),
    limit: int = Query(default=20, ge=1, le=100),
) -> CursorPage[VisitSummary]:
    """Full visit history, paginated, newest first."""
    return await FarmerService(db).list_visits(
        farmer_id, user=user, cursor=cursor, limit=limit
    )


@router.get(
    "/{farmer_id}/livestock-history",
    response_model=list[LivestockProfileResponse],
)
async def farmer_livestock_history(
    farmer_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list[LivestockProfileResponse]:
    """Every livestock snapshot for this farmer, newest first — shows how the
    herd/feed data evolved across visits."""
    return await FarmerService(db).livestock_history(farmer_id, user=user)


@router.get("/{farmer_id}/lead-history", response_model=list[LeadHistoryItem])
async def farmer_lead_history(
    farmer_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list[LeadHistoryItem]:
    """All lead status changes with timestamps and reasons, newest first."""
    return await FarmerService(db).lead_history(farmer_id, user=user)


@router.post("/{farmer_id}/lead-status", response_model=LeadResponse, status_code=201)
async def update_lead_status(
    farmer_id: int,
    body: LeadStatusUpdate,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> LeadResponse:
    """Record a lead status change (Hot/Warm/Cold) with a required reason. One
    row per change — full history is preserved (see /lead-history)."""
    return await FarmerService(db).update_lead_status(farmer_id, body, user=user)
