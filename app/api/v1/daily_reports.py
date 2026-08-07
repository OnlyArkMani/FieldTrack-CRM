"""Daily Status Report (DSR) router -- Module 5.

Route ordering: static paths (/my, /team, /archive) declared before
parameterised paths (/{id}/...) to avoid conflicts.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import (
    CurrentUser,
    get_current_manager,
    get_db,
)
from app.core.exceptions import forbidden, not_found
from app.models.crm import DailyReport, Farmer, Visit, VisitPlan, VisitPlanItem
from app.models.enums import UserRole
from app.models.user import User
from app.schemas.common import CursorPage, decode_cursor, encode_cursor
from app.schemas.crm import DailyReportResponse
from app.services.dsr_service import (
    add_manager_comment,
    business_today,
    generate_dsr,
    get_dsr_with_details,
    submit_dsr,
    visits_export_rows,
)

router = APIRouter(prefix="/daily-reports", tags=["daily-reports"])


# -- Request schemas ----------------------------------------------------------

class DsrSubmitRequest(BaseModel):
    end_of_day_note: str | None = Field(
        default=None,
        max_length=300,
        description="Optional end-of-day summary note (max 300 chars)",
    )


class ManagerCommentRequest(BaseModel):
    comment: str = Field(min_length=1, max_length=1000)


# -- Enriched response schemas ------------------------------------------------

class VisitSummaryItem(BaseModel):
    id: int
    farmer_id: int | None = None
    farmer_name: str
    village: str | None = None
    district: str | None = None
    customer_type: str = "FARMER_MEET"
    purpose: str | None
    check_in_at: Any
    check_out_at: Any
    check_in_lat: float | None = None
    check_in_lng: float | None = None
    lead_status: str | None
    meeting_highlights: str | None = None
    farmer_concerns: str | None = None
    product_interest: str | None = None
    vet_required: bool = False
    vet_cattle_count: int | None = None
    order_bags: int | None = None
    order_value: Decimal | None = None
    breed: str | None = None
    current_brand: str | None = None
    livestock_cattle: int | None = None
    price_per_bag: Decimal | None = None


class OrderSummaryItem(BaseModel):
    id: int
    farmer_name: str
    customer_type: str = "FARMER_MEET"
    bags_count: int
    delivery_date: date
    payment_mode: str | None
    price_per_bag: Decimal | None = None
    total_value: Decimal | None = None


class FollowUpSummaryItem(BaseModel):
    id: int
    farmer_name: str
    scheduled_date: date
    scheduled_time: Any
    purpose: str | None


class DsrDetailResponse(DailyReportResponse):
    manager_comment: str | None = None
    visits: list[VisitSummaryItem] = []
    orders: list[OrderSummaryItem] = []
    follow_ups: list[FollowUpSummaryItem] = []


class TeamDsrItem(BaseModel):
    employee_id: int
    employee_name: str
    team_name: str | None = None
    status: str  # SUBMITTED / DRAFT / MISSING
    check_in_time: Any = None
    check_out_time: Any = None
    visits_planned: int = 0
    visits_completed: int
    orders_captured: int
    hot_leads: int
    warm_leads: int
    cold_leads: int
    follow_ups_scheduled: int = 0
    is_late: bool
    submitted_at: Any = None
    has_manager_comment: bool = False
    report_id: int | None


class ArchiveDsrItem(BaseModel):
    """One employee-day row for the date-range (archive) view."""

    id: int | None = None
    employee_id: int
    employee_name: str
    team_name: str | None = None
    report_date: date
    status: str
    visits_completed: int
    orders_captured: int
    hot_leads: int
    warm_leads: int
    cold_leads: int
    is_late: bool
    submitted_at: Any = None


# -- Employee: own DSR history ------------------------------------------------

@router.get("/my", response_model=list[DailyReportResponse])
async def my_dsr_history(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    month: int = Query(default=None, ge=1, le=12),
    year: int = Query(default=None, ge=2020, le=2099),
) -> list[DailyReportResponse]:
    from sqlalchemy import extract

    q = select(DailyReport).where(DailyReport.employee_id == user.id)
    if year:
        q = q.where(extract("year", DailyReport.report_date) == year)
    if month:
        q = q.where(extract("month", DailyReport.report_date) == month)
    q = q.order_by(DailyReport.report_date.desc())

    rows = (await db.execute(q)).scalars().all()
    return [DailyReportResponse.model_validate(r) for r in rows]


@router.get("/my/{report_date}", response_model=DsrDetailResponse)
async def my_dsr_for_date(
    report_date: date,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> DsrDetailResponse:
    detail = await get_dsr_with_details(db, employee_id=user.id, report_date=report_date)
    if detail is None:
        # No DSR row yet — the report is normally created on attendance END, but
        # the employee may open "Today's Summary" before that (or on a day they
        # haven't ended yet). Generate it on-read so the screen always shows the
        # day's visits/orders/leads that WILL be submitted, and yields a
        # submittable DRAFT row. Idempotent (ON CONFLICT DO UPDATE).
        from app.models.attendance import Attendance

        att_id = (
            await db.execute(
                select(Attendance.id).where(
                    Attendance.user_id == user.id,
                    Attendance.date == report_date,
                )
            )
        ).scalar_one_or_none()
        await generate_dsr(
            employee_id=user.id, attendance_id=att_id, report_date=report_date
        )
        detail = await get_dsr_with_details(
            db, employee_id=user.id, report_date=report_date
        )
    if detail is None:
        raise HTTPException(
            status_code=404,
            detail="No attendance recorded for this date.",
        )
    return _build_detail_response(detail)


@router.get("/my/{report_date}/download")
async def download_my_dsr(
    report_date: date,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    fmt: str = Query(default="pdf", pattern="^(pdf|csv)$", alias="format"),
) -> StreamingResponse:
    """Download the caller's DSR for one day as PDF or CSV (?format=pdf|csv)."""
    detail = await get_dsr_with_details(db, employee_id=user.id, report_date=report_date)
    if detail is None:
        raise HTTPException(status_code=404, detail="No DSR for this date.")
    if fmt == "csv":
        return _dsr_csv_response(detail, user.name, report_date)
    return _dsr_pdf_response(detail, user.name, report_date)


# -- Employee: submit DSR -----------------------------------------------------

@router.post("/{report_id}/submit", response_model=DailyReportResponse)
async def submit(
    report_id: int,
    body: DsrSubmitRequest,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> DailyReportResponse:
    report = await submit_dsr(
        db,
        report_id=report_id,
        employee_id=user.id,
        end_of_day_note=body.end_of_day_note,
    )
    return DailyReportResponse.model_validate(report)


# -- Manager: team DSRs ----------------------------------------------------

@router.get("/team", response_model=list[TeamDsrItem])
async def team_dsrs(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    report_date: date = Query(default_factory=business_today),
    team_id: int | None = Query(default=None),
) -> list[TeamDsrItem]:
    """DSR status for a team on a date.

    ADMIN: all employees (optionally filtered by ``team_id``).
    MANAGER: always their own team (``team_id`` is ignored).
    Every active employee in scope is listed — those without a DSR row for the
    date show as MISSING, so the table is never empty.
    """
    emp_filters = [User.role == UserRole.EMPLOYEE, User.is_active.is_(True)]
    if user.role == UserRole.MANAGER:
        if not user.team_id:
            return []
        emp_filters.append(User.team_id == user.team_id)
    elif user.role == UserRole.ADMIN:
        if team_id is not None:
            emp_filters.append(User.team_id == team_id)
    else:
        raise forbidden("Not permitted")

    employees = (
        await db.execute(select(User).where(*emp_filters).order_by(User.name.asc()))
    ).scalars().all()
    if not employees:
        return []

    emp_ids = [e.id for e in employees]
    team_names = await _team_name_map(db, {e.team_id for e in employees if e.team_id})

    dsr_q = await db.execute(
        select(DailyReport).where(
            DailyReport.employee_id.in_(emp_ids),
            DailyReport.report_date == report_date,
        )
    )
    dsrs_by_emp: dict[int, DailyReport] = {d.employee_id: d for d in dsr_q.scalars().all()}

    times_by_emp = await _attendance_times(db, emp_ids, report_date)

    items: list[TeamDsrItem] = []
    for emp in employees:
        d = dsrs_by_emp.get(emp.id)
        ci, co = times_by_emp.get(emp.id, (None, None))
        items.append(
            TeamDsrItem(
                employee_id=emp.id,
                employee_name=emp.name,
                team_name=team_names.get(emp.team_id),
                status=d.status if d else "MISSING",
                check_in_time=ci,
                check_out_time=co,
                visits_planned=d.visits_planned if d else 0,
                visits_completed=d.visits_completed if d else 0,
                orders_captured=d.orders_captured if d else 0,
                hot_leads=d.hot_leads if d else 0,
                warm_leads=d.warm_leads if d else 0,
                cold_leads=d.cold_leads if d else 0,
                follow_ups_scheduled=d.follow_ups_scheduled if d else 0,
                is_late=d.is_late if d else False,
                submitted_at=d.submitted_at if d else None,
                has_manager_comment=bool(getattr(d, "manager_comment", None)) if d else False,
                report_id=d.id if d else None,
            )
        )
    return items


@router.get("/team/{employee_id}/{report_date}", response_model=DsrDetailResponse)
async def team_dsr_detail(
    employee_id: int,
    report_date: date,
    manager: Annotated[User, Depends(get_current_manager)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> DsrDetailResponse:
    if manager.role == UserRole.MANAGER:
        emp = await db.get(User, employee_id)
        if emp is None or emp.team_id != manager.team_id:
            raise forbidden("Employee is not on your team")

    detail = await get_dsr_with_details(db, employee_id=employee_id, report_date=report_date)
    if detail is None:
        raise not_found("No DSR found for this employee on this date")
    return _build_detail_response(detail)


@router.get("/team/{employee_id}/{report_date}/download")
async def download_team_dsr(
    employee_id: int,
    report_date: date,
    manager: Annotated[User, Depends(get_current_manager)],
    db: Annotated[AsyncSession, Depends(get_db)],
    fmt: str = Query(default="pdf", pattern="^(pdf|csv)$", alias="format"),
) -> StreamingResponse:
    """Download one team member's DSR for a day as PDF or CSV (?format=pdf|csv)."""
    emp = await db.get(User, employee_id)
    if emp is None:
        raise not_found("Employee not found")
    if manager.role == UserRole.MANAGER and emp.team_id != manager.team_id:
        raise forbidden("Employee is not on your team")
    detail = await get_dsr_with_details(db, employee_id=employee_id, report_date=report_date)
    if detail is None:
        raise not_found("No DSR found for this employee on this date")
    if fmt == "csv":
        return _dsr_csv_response(detail, emp.name, report_date)
    return _dsr_pdf_response(detail, emp.name, report_date)


# -- Manager: add manager comment ------------------------------------------

@router.post("/{report_id}/manager-comment", response_model=DailyReportResponse)
async def post_manager_comment(
    report_id: int,
    body: ManagerCommentRequest,
    manager: Annotated[User, Depends(get_current_manager)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> DailyReportResponse:
    report = await add_manager_comment(
        db,
        report_id=report_id,
        manager_id=manager.id,
        comment=body.comment,
    )
    return DailyReportResponse.model_validate(report)


# -- Admin: archive -----------------------------------------------------------

_ARCHIVE_MAX_RANGE_DAYS = 731  # ~24 months (checklist #55) — was mistakenly
# capped at 31 days, a limit that belongs to the unrelated /reports/generate
# endpoint (kept fast for ad-hoc report generation), not this archive.


@router.get("/archive", response_model=CursorPage[ArchiveDsrItem])
async def archive(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    employee_id: int | None = Query(default=None),
    team_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    status: str | None = Query(default=None),
    search: str | None = Query(
        default=None, description="Customer/farmer name — matches any visit that day"
    ),
    executive_name: str | None = Query(
        default=None, description="Executive/employee name — free-text substring match (checklist #55)"
    ),
    cursor: str | None = Query(default=None),
    limit: int = Query(default=30, ge=1, le=100),
) -> CursorPage[ArchiveDsrItem]:
    """Date-range DSR archive (one row per employee-day), scope-aware.
    Searchable by date range, executive name (`executive_name`), or
    customer/farmer name (`search`) — checklist #55.

    ADMIN: all employees, optionally filtered by team_id / employee_id.
    MANAGER: always restricted to their own team.
    Max ~24-month window (400 otherwise)."""
    from sqlalchemy import exists, func

    if date_from and date_to:
        if date_from > date_to:
            raise HTTPException(status_code=400, detail="date_from must be before date_to")
        if (date_to - date_from).days > _ARCHIVE_MAX_RANGE_DAYS:
            raise HTTPException(
                status_code=400,
                detail="Date range cannot exceed 24 months.",
                headers={"X-Error-Code": "DATE_RANGE_TOO_LARGE"},
            )

    if user.role not in (UserRole.ADMIN, UserRole.MANAGER):
        raise forbidden("Not permitted")

    base_filters = []
    # Scope: managers are pinned to their own team; admins may filter.
    if user.role == UserRole.MANAGER:
        if not user.team_id:
            return CursorPage[ArchiveDsrItem](items=[], next_cursor=None, total=0, has_more=False)
        base_filters.append(User.team_id == user.team_id)
        if employee_id:
            base_filters.append(DailyReport.employee_id == employee_id)
    else:
        if team_id is not None:
            base_filters.append(User.team_id == team_id)
        if employee_id:
            base_filters.append(DailyReport.employee_id == employee_id)

    if date_from:
        base_filters.append(DailyReport.report_date >= date_from)
    if date_to:
        base_filters.append(DailyReport.report_date <= date_to)
    if status:
        base_filters.append(DailyReport.status == status.upper())
    if executive_name:
        # User is already joined into the base query below — a direct filter,
        # no subquery needed (checklist #55: search by executive name).
        base_filters.append(User.name.ilike(f"%{executive_name.strip()}%"))
    if search:
        # A DSR row has no farmer of its own (it's per employee-day) — match
        # if any visit that employee logged that day was with a matching farmer.
        # (Simple UTC-date comparison — same precision the rest of this
        # archive already works at; not worth a business-timezone-aware
        # bucketing here.)
        base_filters.append(
            exists().where(
                Visit.employee_id == DailyReport.employee_id,
                func.date(Visit.check_in_at) == DailyReport.report_date,
                Visit.farmer_id == Farmer.id,
                Farmer.name.ilike(f"%{search.strip()}%"),
            )
        )

    total = (
        await db.execute(
            select(func.count(DailyReport.id))
            .join(User, User.id == DailyReport.employee_id)
            .where(*base_filters)
        )
    ).scalar_one()

    q = (
        select(DailyReport, User.name, User.team_id)
        .join(User, User.id == DailyReport.employee_id)
        .where(*base_filters)
    )
    cursor_id = decode_cursor(cursor)
    if cursor_id:
        q = q.where(DailyReport.id < cursor_id)
    q = q.order_by(DailyReport.id.desc()).limit(limit + 1)
    rows = (await db.execute(q)).all()

    has_more = len(rows) > limit
    page = rows[:limit]
    next_cursor = encode_cursor(page[-1][0].id) if has_more and page else None

    team_names = await _team_name_map(db, {r[2] for r in page if r[2]})
    items = [
        ArchiveDsrItem(
            id=r[0].id,
            employee_id=r[0].employee_id,
            employee_name=r[1],
            team_name=team_names.get(r[2]),
            report_date=r[0].report_date,
            status=r[0].status,
            visits_completed=r[0].visits_completed,
            orders_captured=r[0].orders_captured,
            hot_leads=r[0].hot_leads,
            warm_leads=r[0].warm_leads,
            cold_leads=r[0].cold_leads,
            is_late=r[0].is_late,
            submitted_at=r[0].submitted_at,
        )
        for r in page
    ]
    return CursorPage[ArchiveDsrItem](
        items=items, next_cursor=next_cursor, total=total, has_more=has_more
    )


# -- Granular per-visit Excel export (checklist: each visit detail) -----------

@router.get("/visits-export")
async def visits_export(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date = Query(default_factory=business_today),
    date_to: date = Query(default_factory=business_today),
    team_id: int | None = Query(default=None),
    employee_id: int | None = Query(default=None),
) -> StreamingResponse:
    """One row per completed visit across the caller's scope, as .xlsx. ADMIN:
    all (optional team/employee filter); MANAGER: own team. Includes meeting,
    order, vet and livestock detail so managers get every visit's full picture."""
    rows = await visits_export_rows(
        db,
        user=user,
        date_from=date_from,
        date_to=date_to,
        team_id=team_id,
        employee_id=employee_id,
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Visits"
    headers = [
        "Date", "Employee", "Team", "Customer", "Type", "Village", "District",
        "Check-in", "Check-out", "Lat", "Lng", "Duration (min)", "Loc. warning",
        "Lead", "Order bags", "Order value", "Vet needed", "Vet cattle", "Breed",
        "Current brand", "Cattle", "Price/bag", "Highlights", "Concerns",
        "Product interest",
    ]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="F5A623")
    for c in ws[1]:
        c.font = Font(bold=True, color="1A1A2E")
        c.fill = head_fill

    def _t(dt):
        return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

    for r in rows:
        ws.append([
            r["date"].isoformat() if r["date"] else "",
            r["employee"], r["team"], r["customer"], r["type"], r["village"],
            r["district"], _t(r["check_in"]), _t(r["check_out"]),
            r["lat"] if r["lat"] is not None else "",
            r["lng"] if r["lng"] is not None else "",
            r["duration_min"] if r["duration_min"] is not None else "",
            r["location_warning"], r["lead"], r["order_bags"],
            float(r["order_value"]) if r["order_value"] is not None else "",
            r["vet_needed"],
            r["vet_cattle"] if r["vet_cattle"] is not None else "",
            r["breed"], r["current_brand"],
            r["cattle"] if r["cattle"] is not None else "",
            float(r["price_per_bag"]) if r["price_per_bag"] is not None else "",
            r["highlights"], r["concerns"], r["product_interest"],
        ])

    for i, _h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"visits_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# -- Scope/enrichment helpers -------------------------------------------------

async def _team_name_map(db: AsyncSession, team_ids: set[int]) -> dict[int, str]:
    if not team_ids:
        return {}
    from app.models.user import Team

    rows = await db.execute(select(Team.id, Team.name).where(Team.id.in_(team_ids)))
    return {tid: name for tid, name in rows.all()}


async def _attendance_times(
    db: AsyncSession, emp_ids: list[int], report_date: date
) -> dict[int, tuple[Any, Any]]:
    """First START and last END session timestamp per employee for the date."""
    from sqlalchemy.orm import selectinload

    from app.models.attendance import Attendance
    from app.models.enums import SessionType

    if not emp_ids:
        return {}
    rows = (
        await db.execute(
            select(Attendance)
            .where(
                Attendance.user_id.in_(emp_ids),
                Attendance.date == report_date,
            )
            .options(selectinload(Attendance.sessions))
        )
    ).scalars().all()
    out: dict[int, tuple[Any, Any]] = {}
    for att in rows:
        check_in = check_out = None
        for s in sorted(att.sessions, key=lambda x: x.timestamp):
            if s.type in (SessionType.START, SessionType.RE_CHECKIN) and check_in is None:
                check_in = s.timestamp
            if s.type == SessionType.END:
                check_out = s.timestamp
        out[att.user_id] = (check_in, check_out)
    return out


# -- Internal helpers ---------------------------------------------------------

def _dsr_csv_response(
    detail: dict, employee_name: str, report_date: date
) -> StreamingResponse:
    """Render one employee-day DSR as CSV. Applies all the same presentation
    rules as the PDF version: time-only for check-in/out, 'Expected Delivery
    Date', NA row for empty follow-ups, Remarks column on visits."""
    from zoneinfo import ZoneInfo

    from app.core.config import get_settings

    settings = get_settings()
    try:
        _tz = ZoneInfo(settings.business_timezone)
    except Exception:  # noqa: BLE001
        _tz = ZoneInfo("UTC")

    def _fmt_time(val: Any) -> str:
        """UTC datetime → local HH:MM string."""
        if not val:
            return ""
        if hasattr(val, "astimezone"):
            return val.astimezone(_tz).strftime("%H:%M")
        if isinstance(val, str):
            try:
                dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
                return dt.astimezone(_tz).strftime("%H:%M")
            except Exception:  # noqa: BLE001
                return val
        return str(val)

    def _fmt_date(val: Any) -> str:
        if not val:
            return ""
        if hasattr(val, "isoformat"):
            return val.isoformat()
        return str(val)

    def _remarks(v: dict) -> str:
        parts = [
            (v.get("meeting_highlights") or "").strip(),
            (v.get("farmer_concerns") or "").strip(),
        ]
        return " / ".join(p for p in parts if p)

    report = detail["report"]
    checkpoints = detail.get("checkpoints") or {}
    check_in = checkpoints.get("check_in_at") or getattr(report, "check_in_at", None)
    check_out = checkpoints.get("check_out_at") or getattr(report, "check_out_at", None)

    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Daily Status Report"])
    w.writerow(["Employee", employee_name])
    w.writerow(["Date", report_date.isoformat()])
    w.writerow(["Status", getattr(report, "status", "")])
    if check_in:
        w.writerow(["Check-in", _fmt_time(check_in)])
    if check_out:
        w.writerow(["Check-out", _fmt_time(check_out)])
    w.writerow(["Visits planned", getattr(report, "visits_planned", 0)])
    w.writerow(["Visits completed", getattr(report, "visits_completed", 0)])
    w.writerow(["Orders captured", getattr(report, "orders_captured", 0)])
    w.writerow([
        "Leads (Hot/Warm/Cold)",
        f"{getattr(report, 'hot_leads', 0)}/"
        f"{getattr(report, 'warm_leads', 0)}/"
        f"{getattr(report, 'cold_leads', 0)}",
    ])
    w.writerow(["End-of-day note", getattr(report, "end_of_day_note", "") or ""])
    w.writerow([])

    w.writerow(["Visits"])
    w.writerow(["Customer", "Type", "Purpose", "Check-in", "Check-out", "Lead", "Remarks"])
    for v in detail["visits"]:
        w.writerow([
            v["farmer_name"],
            v.get("customer_type", ""),
            v.get("purpose") or "",
            _fmt_time(v.get("check_in_at")),
            _fmt_time(v.get("check_out_at")),
            v.get("lead_status") or "",
            _remarks(v),
        ])
    w.writerow([])

    w.writerow(["Orders"])
    w.writerow(["Customer", "Type", "Bags", "Expected Delivery Date", "Payment"])
    for o in detail["orders"]:
        w.writerow([
            o["farmer_name"],
            o.get("customer_type", ""),
            o.get("bags_count", ""),
            _fmt_date(o.get("delivery_date")),
            o.get("payment_mode") or "",
        ])
    w.writerow([])

    w.writerow(["Follow-ups"])
    w.writerow(["Customer", "Scheduled Date", "Time", "Purpose"])
    if detail["follow_ups"]:
        for f in detail["follow_ups"]:
            w.writerow([
                f["farmer_name"],
                _fmt_date(f.get("scheduled_date")),
                str(f.get("scheduled_time") or ""),
                f.get("purpose") or "",
            ])
    else:
        w.writerow(["NA", "NA", "NA", "NA"])

    buf.seek(0)
    safe_name = employee_name.replace(" ", "_")
    fname = f"DSR_{safe_name}_{report_date.isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _dsr_pdf_response(
    detail: dict, employee_name: str, report_date: date
) -> StreamingResponse:
    """Render one employee-day DSR as a branded PDF (summary + visits + orders +
    follow-ups). Times are shown in the business timezone (HH:MM only)."""
    from zoneinfo import ZoneInfo

    from reportlab.lib import colors as _c
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from app.core.config import get_settings

    settings = get_settings()
    try:
        _tz = ZoneInfo(settings.business_timezone)
    except Exception:  # noqa: BLE001
        _tz = ZoneInfo("UTC")

    # Brand palette
    _AMBER = _c.HexColor("#F5A623")
    _DARK = _c.HexColor("#1A1A2E")
    _CREAM = _c.HexColor("#FFF8E7")
    _PURPLE = _c.HexColor("#8B7FD4")
    _GRAY = _c.HexColor("#6B6B80")

    _base = getSampleStyleSheet()
    _title_sty = ParagraphStyle(
        "DSRTitle", parent=_base["Title"],
        fontSize=18, textColor=_DARK, spaceAfter=3,
    )
    _meta_sty = ParagraphStyle(
        "DSRMeta", parent=_base["Normal"],
        fontSize=10, textColor=_GRAY, spaceAfter=2, leading=14,
    )
    _sec_sty = ParagraphStyle(
        "DSRSec", parent=_base["Heading3"],
        fontSize=11, textColor=_PURPLE, spaceBefore=8, spaceAfter=4,
    )
    _cell_sty = ParagraphStyle(
        "DSRCell", parent=_base["Normal"], fontSize=8, leading=10,
    )
    _hdr_sty = ParagraphStyle(
        "DSRHdr", parent=_base["Normal"],
        fontSize=8, leading=10, textColor=_c.white,
    )

    def _esc(s: Any) -> str:
        return str(s if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _cell(s: Any) -> Paragraph:
        return Paragraph(_esc(s) if s not in (None, "", "—") else "—", _cell_sty)

    def _hdr(s: str) -> Paragraph:
        return Paragraph(f"<b>{_esc(s)}</b>", _hdr_sty)

    def _fmt_time(val: Any) -> str:
        """UTC datetime → local HH:MM (time only)."""
        if not val:
            return "—"
        if hasattr(val, "astimezone"):
            return val.astimezone(_tz).strftime("%H:%M")
        if isinstance(val, str):
            try:
                dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
                return dt.astimezone(_tz).strftime("%H:%M")
            except Exception:  # noqa: BLE001
                return val
        return str(val)

    def _fmt_date(val: Any) -> str:
        if not val:
            return "—"
        if hasattr(val, "strftime"):
            return val.strftime("%d %b %Y")
        return str(val)

    def _remarks(v: dict) -> str:
        parts = [
            (v.get("meeting_highlights") or "").strip(),
            (v.get("farmer_concerns") or "").strip(),
        ]
        return " / ".join(p for p in parts if p) or "—"

    def _make_table(
        headers: list[str], rows: list[list], col_widths_mm: list[float]
    ) -> Table:
        body: list = [[_hdr(h) for h in headers]]
        for r in rows:
            body.append([_cell(c) for c in r])
        tbl = Table(body, colWidths=[w * mm for w in col_widths_mm], repeatRows=1)
        cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), _AMBER),
            ("TEXTCOLOR", (0, 0), (-1, 0), _c.white),
            ("GRID", (0, 0), (-1, -1), 0.4, _c.HexColor("#D8D4C4")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
        for i in range(1, len(body)):
            if i % 2 == 0:
                cmds.append(("BACKGROUND", (0, i), (-1, i), _CREAM))
        tbl.setStyle(TableStyle(cmds))
        return tbl

    report = detail["report"]
    checkpoints = detail.get("checkpoints") or {}
    check_in = checkpoints.get("check_in_at") or getattr(report, "check_in_at", None)
    check_out = checkpoints.get("check_out_at") or getattr(report, "check_out_at", None)

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        topMargin=15 * mm, bottomMargin=18 * mm,
        leftMargin=15 * mm, rightMargin=15 * mm,
        title="Daily Status Report",
        author="Samarth Sathi",
    )

    elements: list = [
        Paragraph("Daily Status Report", _title_sty),
        Paragraph(
            f"<b>Employee:</b> {_esc(employee_name)}  |  "
            f"<b>Date:</b> {report_date.strftime('%d %B %Y')}",
            _meta_sty,
        ),
        Paragraph(
            f"<b>Status:</b> {_esc(getattr(report, 'status', ''))}  |  "
            f"<b>Check-in:</b> {_fmt_time(check_in)}  |  "
            f"<b>Check-out:</b> {_fmt_time(check_out)}  |  "
            f"<b>Late:</b> {'Yes' if getattr(report, 'is_late', False) else 'No'}",
            _meta_sty,
        ),
        Spacer(1, 4 * mm),
    ]

    # Summary stats in a 4-column label/value grid
    summary_pairs = [
        ("Visits Planned", str(getattr(report, "visits_planned", 0))),
        ("Visits Completed", str(getattr(report, "visits_completed", 0))),
        ("Orders Captured", str(getattr(report, "orders_captured", 0))),
        ("Follow-ups Scheduled", str(getattr(report, "follow_ups_scheduled", 0))),
        ("Hot Leads", str(getattr(report, "hot_leads", 0))),
        ("Warm Leads", str(getattr(report, "warm_leads", 0))),
        ("Cold Leads", str(getattr(report, "cold_leads", 0))),
        ("End-of-day Note", str(getattr(report, "end_of_day_note", "") or "—")),
    ]
    sum_body: list = []
    for i in range(0, len(summary_pairs), 2):
        row: list = []
        for lbl, val in summary_pairs[i : i + 2]:
            row.extend([
                Paragraph(f"<b>{_esc(lbl)}</b>", _cell_sty),
                Paragraph(_esc(val), _cell_sty),
            ])
        while len(row) < 4:
            row.append(Paragraph("", _cell_sty))
        sum_body.append(row)
    sum_tbl = Table(sum_body, colWidths=[48 * mm, 42 * mm, 48 * mm, 42 * mm])
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _CREAM),
        ("BOX", (0, 0), (-1, -1), 0.5, _AMBER),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, _c.HexColor("#E0DCC8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(sum_tbl)
    elements.append(Spacer(1, 4 * mm))

    # Visits — with Remarks column (meeting highlights + farmer concerns)
    # Column widths total 180mm (A4 portrait with 15mm side margins each)
    elements.append(Paragraph("Visits", _sec_sty))
    visit_rows = [
        [
            v["farmer_name"],
            v.get("customer_type", ""),
            v.get("purpose") or "—",
            _fmt_time(v.get("check_in_at")),
            _fmt_time(v.get("check_out_at")),
            v.get("lead_status") or "—",
            _remarks(v),
        ]
        for v in detail["visits"]
    ] or [["—", "—", "—", "—", "—", "—", "—"]]
    elements.append(_make_table(
        ["Customer", "Type", "Purpose", "Check-in", "Check-out", "Lead", "Remarks"],
        visit_rows,
        [38, 18, 22, 14, 14, 18, 56],
    ))
    elements.append(Spacer(1, 4 * mm))

    # Orders — "Expected Delivery Date" label
    elements.append(Paragraph("Orders", _sec_sty))
    order_rows = [
        [
            o["farmer_name"],
            o.get("customer_type", ""),
            str(o.get("bags_count", "")),
            _fmt_date(o.get("delivery_date")),
            o.get("payment_mode") or "—",
        ]
        for o in detail["orders"]
    ] or [["—", "—", "—", "—", "—"]]
    elements.append(_make_table(
        ["Customer", "Type", "Bags", "Expected Delivery Date", "Payment Mode"],
        order_rows,
        [50, 22, 15, 48, 45],
    ))
    elements.append(Spacer(1, 4 * mm))

    # Follow-ups — "NA" row when empty
    elements.append(Paragraph("Follow-ups", _sec_sty))
    fu_rows = [
        [
            f["farmer_name"],
            _fmt_date(f.get("scheduled_date")),
            str(f.get("scheduled_time") or "—"),
            f.get("purpose") or "—",
        ]
        for f in detail["follow_ups"]
    ] or [["NA", "NA", "NA", "NA"]]
    elements.append(_make_table(
        ["Customer", "Scheduled Date", "Time", "Purpose"],
        fu_rows,
        [55, 35, 25, 65],
    ))

    def _footer(canvas, _doc) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(_GRAY)
        canvas.drawString(15 * mm, 8 * mm, "Samarth Sathi — Daily Status Report")
        canvas.drawRightString(195 * mm, 8 * mm, f"Page {_doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)

    safe_name = employee_name.replace(" ", "_")
    fname = f"DSR_{safe_name}_{report_date.isoformat()}.pdf"
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _build_detail_response(detail: dict) -> DsrDetailResponse:
    report = detail["report"]
    base = DailyReportResponse.model_validate(report)
    # NOTE: manager_comment is already a field on DailyReportResponse (added
    # in migration 0006), so it's already in base.model_dump() — passing it
    # again as an explicit kwarg here previously caused a guaranteed
    # `TypeError: got multiple values for keyword argument 'manager_comment'`
    # on every single call to this endpoint. Pre-existing bug, unrelated to
    # this change — found via live testing, fixed here.
    data = base.model_dump()
    # Live checkpoints (checklist #46) win over the daily_reports snapshot,
    # which can go stale — but only when a live value is actually available,
    # so an old report with no attendance row left doesn't lose its data.
    data.update(
        {k: v for k, v in detail.get("checkpoints", {}).items() if v is not None}
    )
    return DsrDetailResponse(
        **data,
        visits=[VisitSummaryItem(**v) for v in detail["visits"]],
        orders=[OrderSummaryItem(**o) for o in detail["orders"]],
        follow_ups=[FollowUpSummaryItem(**f) for f in detail["follow_ups"]],
    )
